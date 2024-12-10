import openpyxl

def parametros(entrada):
    '''
    Función que lee los parámetros de entrada del archivo de enrtada Excel

    :param entrada: archivo de entrada Excel
    :return: lista con los parámetros
    '''
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(entrada)
    sheet1 = workbook['Parametros']
    #   Creacion de la lista donde se almacenan los parámetros
    parametros=list()
    # Leemos la segunda columna las primeras 4 filas
    for i in range(1, 5):
        cell = sheet1.cell(row=i, column=2)
        parametros.append(cell.value)
    return parametros

def generadores(entrada):
    '''
    Función que lee los generadores del archivo de entrada Excel

    :param entrada: archivo de entrada Excel
    :return: listas con los generadores
    '''
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(entrada)
    sheet2 = workbook['Generadores SADI']
    # Creo las listas donde se almacenan los datos
    bus=list()
    governor=list()
    CON=list()
    porcentaje=list()
    idg=list()
    comentario=list()
    tipo=list()

    i=2
    # Lee el 1º bloque de datos
    while sheet2.cell(row=i, column=1).value != None: 
        bus.append(sheet2.cell(row=i, column=1).value)
        governor.append(sheet2.cell(row=i, column=2).value)
        CON.append(sheet2.cell(row=i, column=3).value)
        porcentaje.append(sheet2.cell(row=i, column=4).value)
        idg.append(sheet2.cell(row=i, column=5).value)
        comentario.append(sheet2.cell(row=i, column=6).value)
        tipo.append(sheet2.cell(row=i, column=7).value)
        i+=1
    return bus,governor,CON,porcentaje,idg,comentario,tipo

def regiones_paises_limitrofes(entrada):
    '''
    Función que lee las regiones de los paises limitrofes del archivo de entrada Excel

    :param entrada: archivo de entrada Excel
    :return: listas con las regiones de los paises limitrofes
    '''
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(entrada)
    sheet3 = workbook['Regiones paises limitrofes']
    # Creo las listas donde se almacenan los datos
    n_area=list()
    comment=list()
    i=2
    #Lee el 2º bloque de datos
    while sheet3.cell(row=i, column=1).value != None:
        n_area.append(sheet3.cell(row=i, column=1).value)
        comment.append(sheet3.cell(row=i, column=2).value)
        i+=1
    return n_area,comment

def generadores_no_suman(entrada):
    '''
    Función que lee los generadores que no suman del archivo

    :param entrada: archivo de entrada Excel
    :return: listas con los generadores que no suman
    '''
    # Cargar el archivo de Excel
    workbook = openpyxl.load_workbook(entrada)  
    sheet4 = workbook['Generadores que no suman']
    ibus=list()
    nombre=list()
    id=list()
    i=2
    #Lee el 3º bloque de datos
    while sheet4.cell(row=i, column=1).value != None:
        ibus.append(sheet4.cell(row=i, column=1).value)
        nombre.append(sheet4.cell(row=i, column=2).value)
        id.append(sheet4.cell(row=i, column=3).value)
        i+=1
    return ibus,nombre,id
