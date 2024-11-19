import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def crear (ruta):
    """Crea el archivo excel de salida para los reportes en la ruta especificada
    :param ruta: ruta donde se guardará el archivo excel"""
    # Creamos el excel en la ruta especificada por el parametro
    wb = openpyxl.Workbook()
    # Eliminamos la hoja por defecto
    wb.remove(wb.active)
    # Creamos las hojas del excel
    wb.create_sheet('reserva_total.prn')
    wb.create_sheet('Pmax_Pgen.prn')
    wb.create_sheet('Mayor_maxima.prn')
    wb.create_sheet('Menor_optima.prn')
    wb.create_sheet('Reserva.rep')
    wb.create_sheet('Reserva.err')
    # Creamos los enacabezados de las hojas
    #wb['reserva_total.prn'].append(['Escenario','Reserva Hidro','Reserva Termica','Reserva Total'])#1
    wb['Pmax_Pgen.prn'].append(['IBUS','NOMBRE','ID','POT. MÁXIMA[MW]','POT. OPERATIVA[MW]','RESERVA[MW]','RESERVA PORCENTAJE[%]','PORCENTAJE DATO[%]','RESERVA ÓPTIMA[%]'])
    wb['Mayor_maxima.prn'].append(['IBUS','NOMBRE','ID','POT. MÁXIMA[MW]','POT. OPERATIVA[MW]','RESERVA[MW]','RESERVA PORCENTAJ[%]','PORCENTAJE DATO[%]','RESERVA ÓPTIMA[%]'])
    wb['Menor_optima.prn'].append(['IBUS','NOMBRE','ID','POT. MÁXIMA[MW]','POT. OPERATIVA[MW]','RESERVA[MW]','RESERVA PORCENTAJE[%]','PORCENTAJE DATO[%]','RESERVA ÓPTIMA[%]'])
    wb['Reserva.rep'].append(['Escenario','Reserva Hidro','Reserva Termica','Reserva Total'])#2
    wb['Reserva.err'].append(['Error'])#3
    # Guardamos el excel con el nombre Reserva_salida.xlsx
    ruta_completa= ruta+'/Reserva_salida1.xlsx'
    wb.save(ruta_completa)
    # Mensaje que avisa que el archivo se ha creado en la ruta especificada
    print('Archivo creado en la ruta: ',ruta_completa)
    return

def reserva_total(ruta,reservahidro,reservatermica,reservahidro_rpf,reservatermica_rpf,
                    pot_hidro,pot_TG,pot_CC,pot_TV,reserva_TV,reserva_CC,reserva_TG,
                    generacion_total,reserva_nueva,reservatotal2):
    """Completa los datos de la hoja reserva_total.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta + '/Reserva_salida1.xlsx')
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['reserva_total.prn']
    # Defino formatos y funciones de formatos
    bordes_titulos = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    def bordes_celdas(bordes,fila):
        for col in range(1, 7):
            cell = sheet.cell(row=fila, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = bordes
    
    def unir_celdas_resultados(fila_inicial, fila_final):
        for row in range(fila_inicial, fila_final+1):
            sheet.merge_cells('A'+str(row)+':C'+str(row))
            sheet.merge_cells('D'+str(row)+':F'+str(row))
        return
    
    def bordes_celdas_total(bordes, inicio, final):
        for i in range(inicio, final+1):
            bordes_celdas(bordes,i)
        return

    #Escribimos los datos en la hoja debajo de los encabezados en la fila 2
    sheet['A1'] = 'Análisis de la Reserva Total'
    sheet.merge_cells('A1:F1')
    sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')
    # sheet['A2'] = 'Escenario'
    sheet['A3'] = 'RESERVA ROTANTE EN MAQUINAS QUE REGULAN'
    sheet.merge_cells('A3:F3')
    sheet['A3'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A4'] = 'RESERVA HIDRO [MW]'
    sheet['D4'] = reservahidro
    sheet['A5'] = 'RESERVA TERMICA [MW]'
    sheet['D5'] = reservatermica
    sheet['A6'] = 'RESERVA TOTAL [MW]'
    sheet['D6'] = reservahidro + reservatermica
    unir_celdas_resultados(4,6)
    sheet['A7'] = 'RESERVA ROTANTE DEL PARQUE REGULANTE [%]'
    sheet.merge_cells('A7:E7')
    sheet['A7'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['F7'] =round(((reservatermica+reservahidro)/generacion_total)*100,2)
    sheet['A8'] = 'RESERVA PROGRAMADA A 50Hz PARA RPF'
    sheet.merge_cells('A8:F8')
    sheet['A9'] = 'RESERVA HIDRO [MW]'
    sheet['D9'] = reservahidro_rpf
    sheet['A10'] = 'RESERVA TÉRMICA [MW]'
    sheet['D10'] = reservatermica_rpf
    sheet['A11'] = 'TOTAL SISTEMA [MW]'
    sheet['D11'] = reservahidro_rpf + reservatermica_rpf
    unir_celdas_resultados(9,11)
    sheet['A12'] = 'RESERVA PARA RPF [%]'
    sheet.merge_cells('A12:E12')
    sheet['F12']=round(((reservatermica_rpf+reservahidro_rpf)/generacion_total)*100,2)
    sheet['A13'] = 'COLABORACIÓN DEL PARQUE HIDRO EN RSF [MW]'
    sheet.merge_cells('A13:E13')
    sheet['F13'] = reservahidro-reservahidro_rpf
    sheet['A14'] = 'COLABORACIÓN DEL PARQUE HIDRO EN RSF [%]'
    sheet.merge_cells('A14:E14')
    sheet['F14'] = round(((reservahidro-reservahidro_rpf)/generacion_total)*100,2)
    sheet['A15'] = 'POTENCIA OPERABLE EN EL PARQUE REGULANTE'
    sheet.merge_cells('A15:F15')
    sheet['A15'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A16'] = 'HIDRO [MW]'
    sheet['D16'] = pot_hidro
    sheet['A17'] = 'TÉRMICA TG-CC [MW]'
    sheet['D17'] = pot_TG + pot_CC
    sheet['A18'] = 'TÉRMICA TV [MW]'
    sheet['D18'] = pot_TV
    sheet['A19'] = 'TOTAL [MW]'
    sheet['D19'] = pot_hidro + pot_TG + pot_CC + pot_TV
    unir_celdas_resultados(16,19)
    sheet['A20'] = 'RESERVA PROGRAMADA EN EL PARQUE REGULANTE'
    sheet.merge_cells('A20:F20')
    sheet['A20'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A21'] = 'HIDRO [MW]'
    sheet['D21'] = reservahidro_rpf
    sheet['A22'] = 'TÉRMICA TG-CC [MW]'
    sheet['D22'] = reserva_CC+reserva_TG
    sheet['A23'] = 'TÉRMICA TV [MW]'
    sheet['D23'] = reserva_TV
    sheet['A24'] = 'TOTAL [MW]'
    sheet['D24'] = reserva_TV + reserva_CC + reserva_TG
    unir_celdas_resultados(21,24)
    sheet['A25'] = 'RESERVA NUEVA [MW]'
    sheet.merge_cells('A25:E25')
    sheet['F25'] = reserva_nueva
    sheet['A26'] = 'RESERVA TOTAL 2 [MW]'
    sheet['F26'] = reservatotal2
    sheet.merge_cells('A26:E26')
   

    bordes_celdas_total(bordes_titulos,1,26)
    workbook.save(ruta + '/Reserva_salida1.xlsx')
    workbook.close()
    return

def reserva_total_recorte(ruta,ajuste, tipo_ajustado,reservahidro,reservatermica,reservahidro_rpf,reservatermica_rpf,
                    pot_hidro,pot_TG,pot_CC,pot_TV,reserva_TV,reserva_CC,reserva_TG,
                    generacion_total,reserva_nueva,reservatotal2):
    """Completa los datos de la hoja reserva_total.prn con los valores despues del recorte
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta + '/Reserva_salida1.xlsx')
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['reserva_total.prn']
    # Defino formatos y funciones de formatos
    bordes_titulos = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    def bordes_celdas(bordes,fila):
        for col in range(1, 7):
            cell = sheet.cell(row=fila, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = bordes
    
    def unir_celdas_resultados(fila_inicial, fila_final):
        for row in range(fila_inicial, fila_final+1):
            sheet.merge_cells('A'+str(row)+':C'+str(row))
            sheet.merge_cells('D'+str(row)+':F'+str(row))
        return
    
    def bordes_celdas_total(bordes, inicio, final):
        for i in range(inicio, final+1):
            bordes_celdas(bordes,i)
        return
    
    bordes_celdas_total(bordes_titulos,3,26)

    sheet['A28'] = 'LUEGO DEL RECORTE DE PONTECIA MAXIMA '
    sheet.merge_cells('A28:F28')
    sheet['A28'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A29'] = 'AJUSTE EN BASE A '+ajuste+' SOBRE '+tipo_ajustado
    sheet.merge_cells('A29:F29')
    sheet['A29'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A30'] = 'RESERVA ROTANTE EN MÁQUINAS QUE REGULAN'
    sheet.merge_cells('A30:F30')
    sheet['A31'] = 'RESERVA HIDRO [MW]'
    sheet['D31'] = reservahidro
    sheet['A32'] = 'RESERVA TERMICA [MW]'
    sheet['D32'] = reservatermica
    sheet['A33'] = 'RESERVA TOTAL [MW]'
    sheet['D33'] = reservahidro + reservatermica
    unir_celdas_resultados(31,33)
    sheet['A34'] = 'RESERVA ROTANTE DEL PARQUE REGULANTE [%]'
    sheet.merge_cells('A34:E34')
    sheet['F34'] =round(((reservatermica+reservahidro)/generacion_total)*100,2)
    sheet['A35'] = 'RESERVA PROGRAMADA A 50Hz PARA RPF'
    sheet.merge_cells('A35:F35')
    sheet['A36'] = 'RESERVA HIDRO [MW]'
    sheet['D36'] = reservahidro_rpf
    sheet['A37'] = 'RESERVA TÉRMICA [MW]'
    sheet['D37'] = reservatermica_rpf
    sheet['A38'] = 'TOTAL SISTEMA [MW]'
    sheet['D38'] = reservahidro_rpf + reservatermica_rpf
    unir_celdas_resultados(36,38)
    sheet['A39'] = 'RESERVA PARA RPF [%]'
    sheet.merge_cells('A39:E39')
    sheet['F39']=round(((reservatermica_rpf+reservahidro_rpf)/generacion_total)*100,2)
    sheet['A40'] = 'COLABORACIÓN DEL PARQUE HIDRO EN RSF [MW]'
    sheet.merge_cells('A40:E40')
    sheet['F40'] = reservahidro-reservahidro_rpf
    sheet['A41'] = 'COLABORACIÓN DEL PARQUE HIDRO EN RSF [%]'
    sheet.merge_cells('A41:E41')
    sheet['F41'] = round(((reservahidro-reservahidro_rpf)/generacion_total)*100,2)
    sheet['A42'] = 'POTENCIA OPERABLE EN EL PARQUE REGULANTE'
    sheet.merge_cells('A42:F42')
    bordes_celdas(bordes_titulos,42)
    sheet['A42'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A43'] = 'HIDRO [MW]'
    sheet['D43'] = pot_hidro
    sheet['A44'] = 'TÉRMICA TG-CC [MW]'
    sheet['D44'] = pot_TG + pot_CC
    sheet['A45'] = 'TÉRMICA TV [MW]'
    sheet['D45'] = pot_TV
    sheet['A46'] = 'TOTAL [MW]'
    sheet['D46'] = pot_hidro + pot_TG + pot_CC + pot_TV
    unir_celdas_resultados(43,46)
    sheet['A47'] = 'RESERVA PROGRAMADA EN EL PARQUE REGULANTE'
    sheet.merge_cells('A47:F47')
    sheet['A47'].alignment = openpyxl.styles.Alignment(horizontal='center')
    sheet['A48'] = 'HIDRO [MW]'
    sheet['D48'] = reserva_TV
    sheet['A49'] = 'TÉRMICA TG-CC [MW]'
    sheet['D49'] = reserva_CC+reserva_TG
    sheet['A50'] = 'TÉRMICA TV [MW]'
    sheet['D50'] = reserva_TV
    sheet['A51'] = 'TOTAL [MW]'
    sheet['D51'] = reserva_TV + reserva_CC + reserva_TG
    unir_celdas_resultados(48,51)
    sheet['A52'] = 'RESERVA NUEVA [MW]'
    sheet.merge_cells('A52:E52')
    sheet['F52'] = reserva_nueva
    sheet['A53'] = 'RESERVA TOTAL 2 [MW]'
    sheet['F53'] = reservatotal2
    sheet.merge_cells('A53:E53')

    bordes_celdas_total(bordes_titulos,28,53)
    workbook.save(ruta + '/Reserva_salida1.xlsx')
    workbook.close()
    
    return

def Pmax_Pgen(ruta,ibus,nombre,id,pot_max,pot_gen,max_gen,reserva,por_dato,resopt):
    """Completa los datos de la hoja Pmax_Pgen.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Verificamos que se pueda abrir el excel
    try:
        workbook = openpyxl.load_workbook(ruta+'/Reserva_salida1.xlsx')
    except:
        print('No se pudo abrir el archivo')
        return
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Pmax_Pgen.prn']
    # Escribimos todos los datos de las listas
    for i in range(len(ibus)):
        row=i+2
        sheet.cell(row=row, column=1).value = ibus[i]
        sheet.cell(row=row, column=2).value = nombre[i]
        sheet.cell(row=row, column=3).value = id[i]
        sheet.cell(row=row, column=4).value = round(pot_max[i],2)
        sheet.cell(row=row, column=5).value = round(pot_gen[i],2)
        sheet.cell(row=row, column=6).value = round(max_gen[i],2)
        sheet.cell(row=row, column=7).value = round(reserva[i],2)
        sheet.cell(row=row, column=8).value = por_dato[i]
        sheet.cell(row=row, column=9).value = resopt

    # Se hace el recuadro de los datos y los bordes de las celdas
    for col in range(1, 10):
        cell = sheet.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True) 
    for row in range(1, len(ibus)+2):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    # Se ajusta la celda para que se vea mejor
    for col in range(1, 10):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Guardamos el archivo de excel
    workbook.save(ruta + '/Reserva_salida1.xlsx')
    workbook.close()
    return

def Mayor_maxima(ruta,ibus,nombre,id,pot_max,pot_gen,max_gen,reserva,por_dato,resopt):
    """Completa los datos de la hoja Mayor_maxima.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    try:
        workbook = openpyxl.load_workbook(ruta+'/Reserva_salida1.xlsx')
    except:
        print('No se pudo abrir el archivo')
        return
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Mayor_maxima.prn']
    # Escribimos todos los datos de las listas donde los generadores tenga una reserva mayor a la maxima
    j=0
    for i in range(len(ibus)):
        if reserva[i]>resopt:
            row=j+2
            sheet.cell(row=row, column=1).value = ibus[i]
            sheet.cell(row=row, column=2).value = nombre[i]
            sheet.cell(row=row, column=3).value = id[i]
            sheet.cell(row=row, column=4).value = round(pot_max[i],2)
            sheet.cell(row=row, column=5).value = round(pot_gen[i],2)
            sheet.cell(row=row, column=6).value = round(max_gen[i],2)
            sheet.cell(row=row, column=7).value = reserva[i]
            sheet.cell(row=row, column=8).value = por_dato[i]
            sheet.cell(row=row, column=9).value = resopt
            j+=1
    # Se hace el recuadro de los datos y los bordes de las celdas
    for col in range(1, 10):
        cell = sheet.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True) 
    for row in range(1, j+2):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    # Se ajusta la celda para que se vea mejor
    for col in range(1, 10):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25


    # Guardamos el archivo de excel
    workbook.save(ruta + '/Reserva_salida1.xlsx')
    workbook.close()
    return

def Menor_optima(ruta,ibus,nombre,id,pot_max,pot_gen,max_gen,reserva,por_dato,resopt):
    """Completa los datos de la hoja Menor_optima.prn
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    try:
        workbook = openpyxl.load_workbook(ruta+'/Reserva_salida1.xlsx')
    except:
        print('No se pudo abrir el archivo')
        return
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Menor_optima.prn']
    # Escribimos todos los datos de las listas donde los generadores tenga una reserva mayor a la maxima
    j=0
    for i in range(len(ibus)):
        if reserva[i]<resopt:
            row=j+2
            sheet.cell(row=row, column=1).value = ibus[i]
            sheet.cell(row=row, column=2).value = nombre[i]
            sheet.cell(row=row, column=3).value = id[i]
            sheet.cell(row=row, column=4).value = round(pot_max[i],2)
            sheet.cell(row=row, column=5).value = round(pot_gen[i],2)
            sheet.cell(row=row, column=6).value = round(max_gen[i],2)
            sheet.cell(row=row, column=7).value = reserva[i]
            sheet.cell(row=row, column=8).value = por_dato[i]
            sheet.cell(row=row, column=9).value = resopt
            j+=1

    # Se hace el recuadro de los datos y los bordes de las celdas
    for col in range(1, 10):
        cell = sheet.cell(row=1, column=col)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True) 
    for row in range(1, j+2):
        for col in range(1, 10):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    # Se ajusta la celda para que se vea mejor
    for col in range(1, 10):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    # Guardamos el archivo de excel
    workbook.save(ruta + '/Reserva_salida1.xlsx')
    workbook.close()
    return

def Reserva_rep(ruta):
    """Completa los datos de la hoja Reserva.rep
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta)
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Reserva.rep']
    #Escribimos los datos en la hoja debajo de los encabezados en la fila 2

    return

def Reserva_err(ruta, error):
    """Completa los datos de la hoja Reserva.err
    :param ruta: ruta donde se encuentra el archivo excel de entrada"""
    #Abrimos el archivo de excel
    workbook = openpyxl.load_workbook(ruta +'/Reserva_salida1.xlsx')
    #Seleccionamos la hoja donde vamos a completar con datos
    sheet = workbook['Reserva.err']
    #Escribimos los datos en la primera celda en blanco
    last_row = sheet.max_row + 1
    sheet.cell(row=last_row, column=1, value=error)
    workbook.save(ruta+'/Reserva_salida1.xlsx')
    workbook.close()

    return

